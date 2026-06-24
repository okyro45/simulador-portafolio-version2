import sqlite3
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import (
    getSampleStyleSheet
)


# ==================================================
# REPORTE EXCEL
# ==================================================

def generar_excel_completo():

    conn = sqlite3.connect(
        "data/portfolio.db"
    )

    wb = Workbook()

    # ======================
    # RESUMEN
    # ======================

    ws = wb.active

    ws.title = "Resumen"

    ws["A1"] = (
        "SIMULADOR PROFESIONAL DE PORTAFOLIOS"
    )

    ws["A1"].font = Font(
        bold=True,
        size=14
    )

    # ======================
    # ACCIONES
    # ======================

    df = pd.read_sql_query(
        "SELECT * FROM acciones",
        conn
    )

    ws_acciones = wb.create_sheet(
        "Acciones"
    )

    if not df.empty:

        ws_acciones.append(
            list(df.columns)
        )

        for fila in df.values.tolist():

            ws_acciones.append(fila)

    # ======================
    # BONOS
    # ======================

    df = pd.read_sql_query(
        "SELECT * FROM bonos",
        conn
    )

    ws_bonos = wb.create_sheet(
        "Bonos"
    )

    if not df.empty:

        ws_bonos.append(
            list(df.columns)
        )

        for fila in df.values.tolist():

            ws_bonos.append(fila)

    # ======================
    # FONDOS INMOBILIARIOS
    # ======================

    df = pd.read_sql_query(
        "SELECT * FROM fondos_inmobiliarios",
        conn
    )

    ws_fi = wb.create_sheet(
        "Fondos_Inmobiliarios"
    )

    if not df.empty:

        ws_fi.append(
            list(df.columns)
        )

        for fila in df.values.tolist():

            ws_fi.append(fila)

    # ======================
    # FONDOS BANCARIOS
    # ======================

    df = pd.read_sql_query(
        "SELECT * FROM fondos_bancarios",
        conn
    )

    ws_fb = wb.create_sheet(
        "Fondos_Bancarios"
    )

    if not df.empty:

        ws_fb.append(
            list(df.columns)
        )

        for fila in df.values.tolist():

            ws_fb.append(fila)

    # ======================
    # DEPOSITOS
    # ======================

    df = pd.read_sql_query(
        "SELECT * FROM depositos_plazo",
        conn
    )

    ws_dep = wb.create_sheet(
        "Depositos_Plazo"
    )

    if not df.empty:

        ws_dep.append(
            list(df.columns)
        )

        for fila in df.values.tolist():

            ws_dep.append(fila)

    conn.close()

    archivo = (
        "exports/Reporte_Ejecutivo.xlsx"
    )

    wb.save(
        archivo
    )

    return archivo


# ==================================================
# REPORTE PDF
# ==================================================

def generar_pdf_ejecutivo():

    conn = sqlite3.connect(
        "data/portfolio.db"
    )

    df_acciones = pd.read_sql_query(
        "SELECT * FROM acciones",
        conn
    )

    df_bonos = pd.read_sql_query(
        "SELECT * FROM bonos",
        conn
    )

    df_fi = pd.read_sql_query(
        "SELECT * FROM fondos_inmobiliarios",
        conn
    )

    df_fb = pd.read_sql_query(
        "SELECT * FROM fondos_bancarios",
        conn
    )

    df_dep = pd.read_sql_query(
        "SELECT * FROM depositos_plazo",
        conn
    )

    conn.close()

    total_acciones = (
        df_acciones["importe_invertido"].sum()
        if not df_acciones.empty
        else 0
    )

    total_bonos = (
        df_bonos["importe_invertido"].sum()
        if not df_bonos.empty
        else 0
    )

    total_fi = (
        df_fi["importe_invertido"].sum()
        if not df_fi.empty
        else 0
    )

    total_fb = (
        df_fb["importe_invertido"].sum()
        if not df_fb.empty
        else 0
    )

    total_dep = (
        df_dep["importe"].sum()
        if not df_dep.empty
        else 0
    )

    valor_total = (
        total_acciones +
        total_bonos +
        total_fi +
        total_fb +
        total_dep
    )

    archivo = (
        "exports/Reporte_Ejecutivo.pdf"
    )

    doc = SimpleDocTemplate(
        archivo
    )

    estilos = (
        getSampleStyleSheet()
    )

    elementos = []

    elementos.append(
        Paragraph(
            "SIMULADOR PROFESIONAL DE PORTAFOLIOS",
            estilos["Title"]
        )
    )

    elementos.append(
        Spacer(1, 12)
    )

    elementos.append(
        Paragraph(
            "Universidad Continental",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            "Maestria en Finanzas",
            estilos["Normal"]
        )
    )

    elementos.append(
        Spacer(1, 12)
    )

    elementos.append(
        Paragraph(
            "Resumen Ejecutivo",
            estilos["Heading2"]
        )
    )

    elementos.append(
        Paragraph(
            f"Valor Total del Portafolio: S/ {valor_total:,.2f}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Acciones: S/ {total_acciones:,.2f}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Bonos: S/ {total_bonos:,.2f}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Fondos Inmobiliarios: S/ {total_fi:,.2f}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Fondos Bancarios: S/ {total_fb:,.2f}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Paragraph(
            f"Depositos a Plazo: S/ {total_dep:,.2f}",
            estilos["Normal"]
        )
    )

    doc.build(
        elementos
    )

    return archivo